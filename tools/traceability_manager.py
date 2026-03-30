#!/usr/bin/env python3
"""
Traceability Manager Tool for EN 50128 Projects

Purpose: Create, maintain, and validate traceability matrices across the
         EN 50128 software development lifecycle.

Usage:
    python3 tools/traceability_manager.py <command> [options]

Commands:
    create       Create new traceability matrix template
    validate     Validate traceability completeness
    query        Query traceability links
    report       Generate traceability report
    import       Import traceability data
    export       Export traceability data
    check-gaps   Detect traceability gaps
    extract      Auto-extract traceability from documents
    visualize    Generate traceability visualizations
    sync         Synchronize CSV/JSON/Markdown formats
    gate-check   Check T1-T15 normative rules for a lifecycle phase gate
    verify-link  Mark a traceability link as VER-verified in a matrix CSV

EN 50128 Compliance:
    - SIL 3-4: 100% traceability MANDATORY (Table A.9, Technique 7)
    - SIL 1-2: Traceability Highly Recommended
    - SIL 0: Traceability Recommended

Matrix Naming Convention:
    CSV files stored in evidence/traceability/ are named using the
    document-ID convention derived from TRACEABILITY.md Section 9 (T1-T15):
        doc<from>_to_doc<to>.csv
    Examples:
        docS1_to_doc6.csv   (T1: S1 system requirements -> SRS [6])
        docS4_to_doc6.csv   (T2: S4 safety requirements -> SRS [6])
        doc6_to_doc9.csv    (T3: SRS [6] -> SAS [9])
        doc9_to_doc6.csv    (T4: SAS [9] -> SRS [6])
        doc10_to_doc9.csv   (T5a: SDS [10] -> SAS [9])
        doc10_to_doc6.csv   (T5b: SDS [10] -> SRS [6])
        doc10_to_doc11.csv  (T5c: SDS [10] -> Interface Specs [11])
        doc15_to_doc10.csv  (T6: Component Design [15] -> SDS [10])
        doc18_to_doc15.csv  (T7: Source Code [18] -> Component Design [15])
        doc7_to_doc6.csv    (T8: Overall Test Spec [7] -> SRS [6])
        doc6_to_doc7.csv    (T9: SRS [6] -> Overall Test Spec [7])
        doc12_to_doc6.csv   (T10a: Integration Test Spec [12] -> SRS/SAS/SDS/IS)
        doc13_to_doc6.csv   (T10b: HW/SW Integration Test Spec [13] -> S1+S2+SRS+SAS+SDS)
        doc16_to_doc15.csv  (T11: Component Test Spec [16] -> Component Design [15])
        doc20_to_doc16.csv  (T12a: Component Test Report [20] -> Component Test Spec [16])
        doc21_to_doc12.csv  (T12b: Integration Test Report [21] -> Integration Test Spec [12])
        doc22_to_doc13.csv  (T12c: HW/SW Test Report [22] -> HW/SW Test Spec [13])
        doc24_to_doc7.csv   (T12d: Overall Test Report [24] -> Overall Test Spec [7])
        doc25_to_doc6.csv   (T13: Validation Report [25] -> SRS [6])

Version: 1.1
Date: 2026-03-30
"""

import argparse
import csv
import io
import json
import os
import re
import sys
from dataclasses import dataclass, asdict, field
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Set, Tuple

# ============================================================================
# Data Models
# ============================================================================

@dataclass
class TraceabilityLink:
    """Represents a single traceability link between two artifacts."""
    
    # Link identifiers
    source_id: str
    source_type: str
    target_id: str
    target_type: str
    
    # Link metadata
    link_type: str = "allocated_to"
    rationale: str = ""
    
    # Verification data
    verified: bool = False
    verified_by: str = ""
    verified_date: str = ""
    
    # Quality metadata
    confidence: str = "high"  # high, medium, low
    source_document: str = ""
    
    # Traceability
    created_date: str = ""
    modified_date: str = ""
    version: str = "1.0"
    
    def __post_init__(self):
        """Initialize timestamps if not provided."""
        if not self.created_date:
            self.created_date = datetime.now().isoformat()
        if not self.modified_date:
            self.modified_date = self.created_date

@dataclass
class TraceabilityMatrix:
    """Represents a complete traceability matrix between two artifact types."""
    
    # Matrix metadata
    name: str
    source_type: str
    target_type: str
    project: str
    
    # Links
    links: List[TraceabilityLink] = field(default_factory=list)
    
    # Validation metadata
    coverage: float = 0.0
    completeness: float = 0.0
    sil_level: int = 0
    
    # Status
    status: str = "draft"  # draft, review, approved, baseline
    approved_by: str = ""
    approved_date: str = ""
    
    # File paths
    csv_path: str = ""
    json_path: str = ""
    markdown_path: str = ""
    
    # Version control
    version: str = "1.0"
    baseline: str = ""
    
    # Timestamps
    created_date: str = ""
    modified_date: str = ""
    
    def __post_init__(self):
        """Initialize timestamps if not provided."""
        if not self.created_date:
            self.created_date = datetime.now().isoformat()
        if not self.modified_date:
            self.modified_date = self.created_date

# ============================================================================
# Traceability Manager Core
# ============================================================================

class TraceabilityManager:
    """Core traceability management functionality."""
    
    def __init__(self, project: str = None, evidence_dir: str = "evidence/traceability"):
        """
        Initialize Traceability Manager.
        
        Args:
            project: Project name (reads from LIFECYCLE_STATE.md if None)
            evidence_dir: Directory for traceability evidence
        """
        self.project = project or self._read_project_name()
        self.evidence_dir = Path(evidence_dir)
        self.evidence_dir.mkdir(parents=True, exist_ok=True)
        
        # Load configuration
        self.config = self._load_config()
        
    def _read_project_name(self) -> str:
        """Read project name from LIFECYCLE_STATE.md."""
        try:
            with open("LIFECYCLE_STATE.md", "r") as f:
                content = f.read()
                match = re.search(r'\|\s*\*\*Project Name\*\*\s*\|\s*(.+?)\s*\|', content)
                if match:
                    return match.group(1).strip()
        except FileNotFoundError:
            pass
        return "UnknownProject"
    
    def _load_config(self) -> Dict:
        """Load traceability configuration."""
        # Default configuration
        config = {
            "sil_thresholds": {
                0: 0.80,  # 80% coverage for SIL 0
                1: 0.90,  # 90% coverage for SIL 1
                2: 0.95,  # 95% coverage for SIL 2
                3: 1.00,  # 100% coverage for SIL 3 (MANDATORY)
                4: 1.00,  # 100% coverage for SIL 4 (MANDATORY)
            },
            "artifact_patterns": {
                "software_requirement": r"(SW-REQ-[A-Z]+-[A-Z0-9]+-\d+|REQ-[A-Z]+-[A-Z0-9]+-\d+)",
                "system_requirement": r"SYS-REQ-[A-Z]+-[A-Z0-9]+-\d+",
                "architecture_component": r"ARCH-[A-Z]+-[A-Z0-9]+-\d+",
                "design_component": r"DES-[A-Z]+-[A-Z0-9]+-\d+",
                "test_case": r"TEST-[A-Z]+-[A-Z0-9]+-\d+",
                "function": r"[a-z_][a-z0-9_]*\(\)",
            }
        }
        
        # Load from .traceability.yaml if exists
        config_path = Path(".traceability.yaml")
        if config_path.exists():
            try:
                import yaml
            except ImportError:
                print(
                    "Warning: pyyaml not installed; .traceability.yaml will not be loaded. "
                    "Install it with: pip install pyyaml",
                    file=sys.stderr,
                )
                return config
            with open(config_path) as f:
                user_config = yaml.safe_load(f)
                config.update(user_config)
        
        return config
    
    # ========================================================================
    # Command: create
    # ========================================================================
    
    def create_matrix(self, source_type: str, target_type: str, output_path: str = None) -> TraceabilityMatrix:
        """
        Create new traceability matrix template.
        
        Args:
            source_type: Source artifact type (e.g., "requirements")
            target_type: Target artifact type (e.g., "architecture")
            output_path: Optional output path (default: evidence/traceability/)
        
        Returns:
            TraceabilityMatrix: Created matrix
        """
        matrix_name = f"{source_type}_to_{target_type}"
        
        # Create matrix object
        matrix = TraceabilityMatrix(
            name=matrix_name,
            source_type=source_type,
            target_type=target_type,
            project=self.project,
        )
        
        # Set file paths
        if output_path:
            base_path = Path(output_path)
        else:
            base_path = self.evidence_dir / matrix_name
        
        matrix.csv_path = str(base_path.with_suffix('.csv'))
        matrix.json_path = str(base_path.with_suffix('.json'))
        matrix.markdown_path = str(base_path.with_suffix('.md'))
        
        # Create empty CSV template
        self._write_csv(matrix)
        
        # Create JSON index
        self._write_json(matrix)
        
        # Create Markdown report
        self._write_markdown(matrix)
        
        print(f"✓ Created traceability matrix: {matrix_name}")
        print(f"  CSV:      {matrix.csv_path}")
        print(f"  JSON:     {matrix.json_path}")
        print(f"  Markdown: {matrix.markdown_path}")
        
        return matrix
    
    # ========================================================================
    # Command: validate
    # ========================================================================
    
    def validate(self, matrix_name: str = None, phase: str = None, sil: int = 0) -> Dict:
        """
        Validate traceability completeness.
        
        Args:
            matrix_name: Specific matrix to validate (e.g., "requirements_to_architecture")
            phase: Validate all matrices for lifecycle phase
            sil: SIL level (0-4) for validation thresholds
        
        Returns:
            Dict with validation results
        """
        results = {
            "project": self.project,
            "sil": sil,
            "threshold": self.config["sil_thresholds"][sil],
            "matrices": [],
            "overall_pass": True,
        }
        
        # Determine which matrices to validate
        if matrix_name:
            matrices_to_validate = [matrix_name]
        elif phase:
            matrices_to_validate = self._get_phase_matrices(phase)
        else:
            # Validate all matrices
            matrices_to_validate = [p.stem for p in self.evidence_dir.glob("*.csv")]
        
        for name in matrices_to_validate:
            matrix_result = self._validate_matrix(name, sil)
            results["matrices"].append(matrix_result)
            
            if not matrix_result["pass"]:
                results["overall_pass"] = False
        
        # Print results
        self._print_validation_results(results)
        
        return results
    
    def _validate_matrix(self, matrix_name: str, sil: int) -> Dict:
        """Validate a single traceability matrix."""
        csv_path = self.evidence_dir / f"{matrix_name}.csv"
        
        if not csv_path.exists():
            return {
                "name": matrix_name,
                "pass": False,
                "error": "Matrix file not found",
                "coverage": 0.0,
                "threshold": self.config["sil_thresholds"][sil],
                "total_sources": 0,
                "traced_sources": 0,
                "untraceable_sources": 0,
            }
        
        # Load matrix
        matrix = self._read_csv(str(csv_path))
        
        # Calculate coverage
        sources = set()
        sources_with_targets = set()
        
        for link in matrix.links:
            sources.add(link.source_id)
            if link.target_id:
                sources_with_targets.add(link.source_id)
        
        coverage = len(sources_with_targets) / len(sources) if sources else 0.0
        threshold = self.config["sil_thresholds"][sil]
        
        return {
            "name": matrix_name,
            "pass": coverage >= threshold,
            "coverage": coverage,
            "threshold": threshold,
            "total_sources": len(sources),
            "traced_sources": len(sources_with_targets),
            "untraceable_sources": len(sources) - len(sources_with_targets),
        }
    
    def _get_phase_matrices(self, phase: str) -> List[str]:
        """Get required traceability matrices for a lifecycle phase.

        Matrix names follow the document-ID convention:
            doc<from>_to_doc<to>
        where <from> and <to> are Annex C item numbers (integers) or
        system document identifiers (S1, S2, S3, S4).

        This convention is derived from the T1-T15 rules in
        activities/traceability.yaml (TRACEABILITY.md Section 9).
        """
        # Rules by phase (cumulative — later phases include all prior matrices
        # that must still be valid at that gate).
        phase_matrices = {
            # Phase 2 (requirements): T1, T2, T8, T9
            "requirements": [
                "docS1_to_doc6",    # T1: S1 -> SRS [6]
                "docS4_to_doc6",    # T2: S4 -> SRS [6]
                "doc7_to_doc6",     # T8: Overall Test Spec [7] -> SRS [6]
                "doc6_to_doc7",     # T9: SRS [6] -> Overall Test Spec [7]
            ],
            # Phase 3 (design): + T3, T4, T5a, T5b, T5c, T10a, T10b
            "design": [
                "docS1_to_doc6",    # T1
                "docS4_to_doc6",    # T2
                "doc6_to_doc9",     # T3: SRS [6] -> SAS [9]
                "doc9_to_doc6",     # T4: SAS [9] -> SRS [6]
                "doc10_to_doc9",    # T5a: SDS [10] -> SAS [9]
                "doc10_to_doc6",    # T5b: SDS [10] -> SRS [6]
                "doc10_to_doc11",   # T5c: SDS [10] -> Interface Specs [11]
                "doc7_to_doc6",     # T8
                "doc6_to_doc7",     # T9
                "doc12_to_doc6",    # T10a: Integration Test Spec [12] -> SRS [6]
                "doc13_to_doc6",    # T10b: HW/SW Integration Test Spec [13] -> SRS [6]
            ],
            # Phase 4 (component-design): + T6, T11
            "component-design": [
                "docS1_to_doc6",
                "docS4_to_doc6",
                "doc6_to_doc9",
                "doc9_to_doc6",
                "doc10_to_doc9",
                "doc10_to_doc6",
                "doc10_to_doc11",
                "doc15_to_doc10",   # T6: Component Design [15] -> SDS [10]
                "doc7_to_doc6",
                "doc6_to_doc7",
                "doc12_to_doc6",
                "doc13_to_doc6",
                "doc16_to_doc15",   # T11: Component Test Spec [16] -> Component Design [15]
            ],
            # Phase 5 (implementation-testing): + T7, T12a
            "implementation-testing": [
                "docS1_to_doc6",
                "docS4_to_doc6",
                "doc6_to_doc9",
                "doc9_to_doc6",
                "doc10_to_doc9",
                "doc10_to_doc6",
                "doc10_to_doc11",
                "doc15_to_doc10",
                "doc18_to_doc15",   # T7: Source Code [18] -> Component Design [15]
                "doc7_to_doc6",
                "doc6_to_doc7",
                "doc12_to_doc6",
                "doc13_to_doc6",
                "doc16_to_doc15",
                "doc20_to_doc16",   # T12a: Component Test Report [20] -> Test Spec [16]
            ],
            # Phase 6 (integration): + T10a/T10b reports, T12b, T12c
            "integration": [
                "docS1_to_doc6",
                "docS4_to_doc6",
                "doc6_to_doc9",
                "doc9_to_doc6",
                "doc10_to_doc9",
                "doc10_to_doc6",
                "doc10_to_doc11",
                "doc15_to_doc10",
                "doc18_to_doc15",
                "doc7_to_doc6",
                "doc6_to_doc7",
                "doc12_to_doc6",
                "doc13_to_doc6",
                "doc16_to_doc15",
                "doc20_to_doc16",
                "doc21_to_doc12",   # T12b: Integration Test Report [21] -> Test Spec [12]
                "doc22_to_doc13",   # T12c: HW/SW Test Report [22] -> HW/SW Test Spec [13]
            ],
            # Phase 7 (validation): + T12d, T13
            "validation": [
                "docS1_to_doc6",
                "docS4_to_doc6",
                "doc6_to_doc9",
                "doc9_to_doc6",
                "doc10_to_doc9",
                "doc10_to_doc6",
                "doc10_to_doc11",
                "doc15_to_doc10",
                "doc18_to_doc15",
                "doc7_to_doc6",
                "doc6_to_doc7",
                "doc12_to_doc6",
                "doc13_to_doc6",
                "doc16_to_doc15",
                "doc20_to_doc16",
                "doc21_to_doc12",
                "doc22_to_doc13",
                "doc24_to_doc7",    # T12d: Overall Test Report [24] -> Overall Test Spec [7]
                "doc25_to_doc6",    # T13: Validation Report [25] -> SRS [6]
            ],
        }

        return phase_matrices.get(phase, [])
    
    def _print_validation_results(self, results: Dict):
        """Print validation results to console."""
        print(f"\n{'='*70}")
        print(f"Traceability Validation Report")
        print(f"{'='*70}")
        print(f"Project:    {results['project']}")
        print(f"SIL Level:  {results['sil']}")
        print(f"Threshold:  {results['threshold']*100:.0f}%")
        print(f"{'='*70}\n")
        
        for matrix in results["matrices"]:
            status = "✓ PASS" if matrix["pass"] else "✗ FAIL"
            print(f"{status}  {matrix['name']}")
            print(f"       Coverage: {matrix['coverage']*100:.1f}% ({matrix['traced_sources']}/{matrix['total_sources']})")
            
            if not matrix["pass"]:
                print(f"       ⚠ {matrix['untraceable_sources']} sources not traced")
            print()
        
        print(f"{'='*70}")
        overall_status = "✓ PASS" if results["overall_pass"] else "✗ FAIL"
        print(f"Overall Status: {overall_status}")
        print(f"{'='*70}\n")
    
    # ========================================================================
    # Command: query
    # ========================================================================
    
    def query(self, source_id: str = None, target_id: str = None, direction: str = "forward") -> List[TraceabilityLink]:
        """
        Query traceability links.
        
        Args:
            source_id: Source artifact ID (e.g., "SW-REQ-015")
            target_id: Target artifact ID (e.g., "ARCH-COMP-005")
            direction: Query direction (forward, backward, both)
        
        Returns:
            List of TraceabilityLink objects
        """
        all_matrices = list(self.evidence_dir.glob("*.csv"))
        results = []
        
        for csv_path in all_matrices:
            matrix = self._read_csv(str(csv_path))
            
            for link in matrix.links:
                if source_id and link.source_id == source_id:
                    if direction in ["forward", "both"]:
                        results.append(link)
                
                if target_id and link.target_id == target_id:
                    if direction in ["backward", "both"]:
                        results.append(link)
        
        # Print results
        self._print_query_results(results, source_id, target_id, direction)
        
        return results
    
    def _print_query_results(self, results: List[TraceabilityLink], source_id: str, target_id: str, direction: str):
        """Print query results to console."""
        print(f"\n{'='*70}")
        print(f"Traceability Query Results")
        print(f"{'='*70}")
        
        if source_id:
            print(f"Source:    {source_id}")
        if target_id:
            print(f"Target:    {target_id}")
        print(f"Direction: {direction}")
        print(f"Found:     {len(results)} links")
        print(f"{'='*70}\n")
        
        for link in results:
            print(f"{link.source_id} → {link.target_id}")
            print(f"  Type:       {link.link_type}")
            print(f"  Rationale:  {link.rationale}")
            print(f"  Verified:   {link.verified} (by {link.verified_by})")
            print()
    
    # ========================================================================
    # Command: check-gaps
    # ========================================================================
    
    def check_gaps(self, phase: str = None, sil: int = 0) -> Dict:
        """
        Detect traceability gaps.
        
        Args:
            phase: Lifecycle phase to check
            sil: SIL level
        
        Returns:
            Dict with gap analysis
        """
        gaps = {
            "orphan_sources": [],      # Sources with no target
            "orphan_targets": [],      # Targets with no source
            "unverified_links": [],    # Links not verified
            "low_confidence": [],      # Auto-extracted links with low confidence
        }
        
        # Determine matrices to check
        if phase:
            matrices = self._get_phase_matrices(phase)
        else:
            matrices = [p.stem for p in self.evidence_dir.glob("*.csv")]
        
        # Collect all sources and targets
        all_sources = set()
        all_targets = set()
        traced_sources = set()
        traced_targets = set()
        
        for matrix_name in matrices:
            csv_path = self.evidence_dir / f"{matrix_name}.csv"
            if not csv_path.exists():
                continue
            
            matrix = self._read_csv(str(csv_path))
            
            for link in matrix.links:
                all_sources.add(link.source_id)
                all_targets.add(link.target_id)
                
                if link.target_id:
                    traced_sources.add(link.source_id)
                    traced_targets.add(link.target_id)
                
                if not link.verified:
                    gaps["unverified_links"].append({
                        "matrix": matrix_name,
                        "link": f"{link.source_id} → {link.target_id}",
                    })
                
                if link.confidence == "low":
                    gaps["low_confidence"].append({
                        "matrix": matrix_name,
                        "link": f"{link.source_id} → {link.target_id}",
                    })
        
        # Identify orphans (exclude empty-string entries that arise from links with no target)
        gaps["orphan_sources"] = list((all_sources - traced_sources) - {""})
        gaps["orphan_targets"] = list((all_targets - traced_targets) - {""})
        
        # Print results
        self._print_gap_analysis(gaps, sil)
        
        return gaps

    # ========================================================================
    # Command: gate-check
    # ========================================================================

    def gate_check(self, phase: str, sil: int) -> Dict:
        """
        Check EN 50128 T1-T15 normative traceability rules for a lifecycle phase gate.

        Reads activities/traceability.yaml to determine which T1-T15 rules apply
        to the requested phase, then verifies that a matrix CSV named according to
        the document-ID convention (doc<from>_to_doc<to>) exists for each rule and
        meets the SIL coverage threshold.

        Args:
            phase: Lifecycle phase name (planning, requirements, design,
                   component-design, implementation-testing, integration, validation)
            sil: SIL level (0-4)

        Returns:
            Dict with per-rule PASS/FAIL results and overall gate status.
            Exits with code 1 if any rule fails (suitable for CI integration).
        """
        # Rules applicable per phase (earliest phase at which the rule is first checked)
        # T15 is meta (VER process obligation) — not a matrix check; included as reminder.
        RULE_FIRST_PHASE = {
            "T1":   "requirements",
            "T2":   "requirements",
            "T8":   "requirements",
            "T9":   "requirements",
            "T3":   "design",
            "T4":   "design",
            "T5a":  "design",
            "T5b":  "design",
            "T5c":  "design",
            "T10a": "design",
            "T10b": "design",
            "T6":   "component-design",
            "T11":  "component-design",
            "T7":   "implementation-testing",
            "T12":  "integration",       # reports T12a..T12d checked progressively
            "T13":  "validation",
            "T14":  "validation",
            "T15":  "requirements",      # process check — not a matrix
        }

        PHASE_ORDER = [
            "planning",
            "requirements",
            "design",
            "component-design",
            "implementation-testing",
            "integration",
            "validation",
        ]

        # Normalise phase name
        phase_norm = phase.lower().replace(" ", "-")
        if phase_norm not in PHASE_ORDER:
            print(f"✗ Unknown phase: {phase}")
            print(f"  Valid phases: {', '.join(PHASE_ORDER)}")
            return {"overall_pass": False, "error": f"Unknown phase: {phase}"}

        phase_idx = PHASE_ORDER.index(phase_norm)

        # Load traceability spec from YAML
        rules_spec = self._load_traceability_rules()

        # Determine which rules apply up to and including this phase
        applicable_rules = [
            rid for rid, first in RULE_FIRST_PHASE.items()
            if first in PHASE_ORDER and PHASE_ORDER.index(first) <= phase_idx
        ]
        # T15 is a process note — skip matrix check for it
        matrix_rules = [rid for rid in applicable_rules if rid != "T15"]

        results = {
            "project": self.project,
            "phase": phase_norm,
            "sil": sil,
            "threshold": self.config["sil_thresholds"][sil],
            "rules": [],
            "process_notes": [],
            "overall_pass": True,
        }

        # Map each rule to its matrix name(s) and check them
        RULE_MATRICES = {
            "T1":   [("docS1_to_doc6",  "S1 → [6]")],
            "T2":   [("docS4_to_doc6",  "S4 → [6]")],
            "T3":   [("doc6_to_doc9",   "[6] → [9]")],
            "T4":   [("doc9_to_doc6",   "[9] → [6]")],
            "T5a":  [("doc10_to_doc9",  "[10] → [9]")],
            "T5b":  [("doc10_to_doc6",  "[10] → [6]")],
            "T5c":  [("doc10_to_doc11", "[10] → [11]")],
            "T6":   [("doc15_to_doc10", "[15] → [10]")],
            "T7":   [("doc18_to_doc15", "[18] → [15]")],
            "T8":   [("doc7_to_doc6",   "[7] → [6]")],
            "T9":   [("doc6_to_doc7",   "[6] → [7]")],
            "T10a": [("doc12_to_doc6",  "[12] → [6]"),
                     ("doc12_to_doc9",  "[12] → [9]"),
                     ("doc12_to_doc10", "[12] → [10]"),
                     ("doc12_to_doc11", "[12] → [11]")],
            "T10b": [("doc13_to_doc6",  "[13] → [6]"),
                     ("doc13_to_doc9",  "[13] → [9]"),
                     ("doc13_to_doc10", "[13] → [10]"),
                     ("doc13_to_docS1", "[13] → S1"),
                     ("doc13_to_docS2", "[13] → S2")],
            "T11":  [("doc16_to_doc15", "[16] → [15]")],
            "T12":  [("doc20_to_doc16", "[20] → [16]"),
                     ("doc21_to_doc12", "[21] → [12]"),
                     ("doc22_to_doc13", "[22] → [13]"),
                     ("doc24_to_doc7",  "[24] → [7]")],
            "T13":  [("doc25_to_doc6",  "[25] → [6]")],
            "T14":  [],   # Covered by SIL threshold in each matrix check
        }

        for rule_id in sorted(matrix_rules, key=lambda r: (RULE_FIRST_PHASE.get(r, "z"), r)):
            rule_spec = rules_spec.get(rule_id, {})
            rule_text = rule_spec.get("rule", "(see TRACEABILITY.md Section 9)").strip()
            clause = rule_spec.get("normative_clause", "")
            scope = rule_spec.get("scope", "")
            matrices = RULE_MATRICES.get(rule_id, [])

            if rule_id == "T14":
                # T14 is satisfied by the SIL thresholds applied in every other rule
                rule_result = {
                    "rule_id": rule_id,
                    "scope": scope,
                    "normative_clause": clause,
                    "rule_text": rule_text,
                    "pass": True,
                    "note": (
                        f"SIL {sil} threshold ({self.config['sil_thresholds'][sil]*100:.0f}%) "
                        "applied to all matrix checks above."
                    ),
                    "matrices": [],
                }
                results["rules"].append(rule_result)
                continue

            matrix_checks = []
            rule_pass = True

            for matrix_name, link_scope in matrices:
                check = self._validate_matrix(matrix_name, sil)
                check["link_scope"] = link_scope
                matrix_checks.append(check)
                if not check["pass"]:
                    rule_pass = False

            rule_result = {
                "rule_id": rule_id,
                "scope": scope,
                "normative_clause": clause,
                "rule_text": rule_text,
                "pass": rule_pass,
                "matrices": matrix_checks,
            }
            results["rules"].append(rule_result)
            if not rule_pass:
                results["overall_pass"] = False

        # T15 process note
        if "T15" in applicable_rules:
            t15_spec = rules_spec.get("T15", {})
            results["process_notes"].append({
                "rule_id": "T15",
                "normative_clause": t15_spec.get("normative_clause", ""),
                "note": (
                    "T15 is a VER process obligation (not a matrix check). "
                    "VER must record traceability completeness findings in the "
                    "Verification Report. COD reads that report and blocks the gate "
                    "if gaps are found. This tool confirms matrix coverage only."
                ),
            })

        self._print_gate_check_results(results)
        return results

    def _load_traceability_rules(self) -> Dict:
        """
        Load T1-T15 rule specifications from activities/traceability.yaml.

        Returns a dict keyed by rule ID (T1, T2, ..., T15) with fields:
            rule, scope, normative_clause, document_chain.

        Falls back to an empty dict if the YAML file is not found or pyyaml
        is not installed, so gate-check still functions (with reduced detail).
        """
        yaml_path = Path("activities/traceability.yaml")
        if not yaml_path.exists():
            return {}

        try:
            import yaml
        except ImportError:
            print(
                "Warning: pyyaml not installed; T1-T15 rule descriptions will be "
                "unavailable (gate-check still functions but shows generic text). "
                "Install it with: pip install pyyaml",
                file=sys.stderr,
            )
            return {}

        with open(yaml_path) as f:
            data = yaml.safe_load(f)

        rules = {}
        for entry in data.get("traceability", {}).get("traceability_rules", []):
            rid = entry.get("id")
            if rid:
                rules[rid] = entry
        return rules

    def _print_gate_check_results(self, results: Dict):
        """Print gate-check results to console."""
        threshold_pct = results["threshold"] * 100
        print(f"\n{'='*72}")
        print(f"EN 50128 Traceability Gate Check — Phase: {results['phase'].upper()}")
        print(f"{'='*72}")
        print(f"Project:   {results['project']}")
        print(f"SIL Level: {results['sil']}")
        print(f"Threshold: {threshold_pct:.0f}%")
        print(f"{'='*72}\n")

        for rule in results["rules"]:
            status = "PASS" if rule["pass"] else "FAIL"
            status_sym = "✓" if rule["pass"] else "✗"
            print(f"{status_sym} {rule['rule_id']:5s}  [{status}]  {rule['scope']}")
            print(f"         {rule['normative_clause']}")

            note = rule.get("note")
            if note:
                print(f"         Note: {note}")

            for m in rule.get("matrices", []):
                m_status = "✓" if m["pass"] else "✗"
                if "error" in m:
                    print(f"           {m_status} {m['name']:45s}  MISSING (matrix CSV not found)")
                else:
                    cov_pct = m["coverage"] * 100
                    print(
                        f"           {m_status} {m['name']:45s}  "
                        f"{cov_pct:.1f}%  ({m['traced_sources']}/{m['total_sources']})"
                    )
            print()

        if results["process_notes"]:
            print(f"{'─'*72}")
            print("Process Notes (not matrix checks):")
            for note in results["process_notes"]:
                print(f"  [{note['rule_id']}]  {note['normative_clause']}")
                print(f"        {note['note']}")
            print()

        print(f"{'='*72}")
        overall_sym = "✓" if results["overall_pass"] else "✗"
        overall_status = "PASS — gate may proceed" if results["overall_pass"] else "FAIL — gate BLOCKED"
        print(f"Overall: {overall_sym} {overall_status}")
        print(f"{'='*72}\n")

    # ========================================================================
    # Command: verify-link
    # ========================================================================

    def verify_link(
        self,
        matrix_name: str,
        source_id: str,
        target_id: str,
        verified_by: str,
        verified_date: str = None,
    ) -> bool:
        """
        Mark a traceability link as VER-verified in a matrix CSV.

        Sets verified=true, verified_by, and verified_date on the matching
        (source_id, target_id) row.  Persists to CSV and regenerates JSON/MD.

        Args:
            matrix_name: Matrix file stem (e.g., "doc6_to_doc9")
            source_id: Source artifact ID to match
            target_id: Target artifact ID to match
            verified_by: VER reviewer identifier (name or role)
            verified_date: ISO date string (defaults to today)

        Returns:
            True if the link was found and updated, False otherwise.
        """
        if verified_date is None:
            verified_date = datetime.now().strftime("%Y-%m-%d")

        csv_path = self.evidence_dir / f"{matrix_name}.csv"
        if not csv_path.exists():
            print(f"✗ Matrix not found: {matrix_name}", file=sys.stderr)
            return False

        matrix = self._read_csv(str(csv_path))
        matched = False

        for link in matrix.links:
            if link.source_id == source_id and link.target_id == target_id:
                link.verified = True
                link.verified_by = verified_by
                link.verified_date = verified_date
                link.modified_date = datetime.now().isoformat()
                matched = True

        if not matched:
            print(
                f"✗ Link not found: {source_id} → {target_id} in {matrix_name}",
                file=sys.stderr,
            )
            return False

        self._write_csv(matrix)
        self._write_json(matrix)
        self._write_markdown(matrix)

        print(f"✓ Verified link: {source_id} → {target_id}")
        print(f"  Matrix:      {matrix_name}")
        print(f"  Verified by: {verified_by}")
        print(f"  Date:        {verified_date}")
        return True

    def _print_gap_analysis(self, gaps: Dict, sil: int):
        """Print gap analysis to console."""
        print(f"\n{'='*70}")
        print(f"Traceability Gap Analysis (SIL {sil})")
        print(f"{'='*70}\n")
        
        print(f"Orphan Sources (no targets): {len(gaps['orphan_sources'])}")
        for source in gaps['orphan_sources'][:10]:  # Show first 10
            print(f"  - {source}")
        if len(gaps['orphan_sources']) > 10:
            print(f"  ... and {len(gaps['orphan_sources']) - 10} more")
        print()
        
        print(f"Orphan Targets (no sources): {len(gaps['orphan_targets'])}")
        for target in gaps['orphan_targets'][:10]:
            print(f"  - {target}")
        if len(gaps['orphan_targets']) > 10:
            print(f"  ... and {len(gaps['orphan_targets']) - 10} more")
        print()
        
        print(f"Unverified Links: {len(gaps['unverified_links'])}")
        if sil >= 3:
            print("  ⚠ WARNING: All links must be verified for SIL 3-4")
        print()
        
        print(f"Low Confidence Links: {len(gaps['low_confidence'])}")
        if gaps['low_confidence']:
            print("  ⚠ Recommend manual review")
        print()
        
        print(f"{'='*70}\n")
    
    # ========================================================================
    # Command: report
    # ========================================================================
    
    def report(self, source_type: str, target_types: List[str], format: str = "markdown", output: str = None) -> str:
        """
        Generate traceability report.
        
        Args:
            source_type: Source artifact type
            target_types: List of target artifact types
            format: Output format (markdown, csv, json)
            output: Output file path
        
        Returns:
            Report content as string
        """
        # Collect data for all requested matrices
        matrix_data = []
        for target_type in target_types:
            matrix_name = f"{source_type}_to_{target_type}"
            csv_path = self.evidence_dir / f"{matrix_name}.csv"
            entry = {"matrix": matrix_name, "source_type": source_type, "target_type": target_type}
            if csv_path.exists():
                matrix = self._read_csv(str(csv_path))
                total_sources = len(set(lnk.source_id for lnk in matrix.links))
                traced_sources = len(set(lnk.source_id for lnk in matrix.links if lnk.target_id))
                entry["coverage"] = traced_sources / total_sources if total_sources > 0 else 0.0
                entry["total_links"] = len(matrix.links)
                entry["links"] = [
                    {
                        "source_id": lnk.source_id,
                        "target_id": lnk.target_id,
                        "link_type": lnk.link_type,
                        "verified": lnk.verified,
                        "confidence": lnk.confidence,
                    }
                    for lnk in sorted(matrix.links, key=lambda x: x.source_id)
                ]
                entry["found"] = True
            else:
                entry["found"] = False
                entry["coverage"] = 0.0
                entry["total_links"] = 0
                entry["links"] = []
            matrix_data.append(entry)

        if format == "json":
            report_content = json.dumps(
                {
                    "project": self.project,
                    "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "source_type": source_type,
                    "matrices": matrix_data,
                },
                indent=2,
            )

        elif format == "csv":
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(["matrix", "source_id", "target_id", "link_type", "verified", "confidence"])
            for entry in matrix_data:
                if not entry["found"]:
                    writer.writerow([entry["matrix"], "", "", "", "", ""])
                    continue
                for lnk in entry["links"]:
                    writer.writerow([
                        entry["matrix"],
                        lnk["source_id"],
                        lnk["target_id"],
                        lnk["link_type"],
                        "yes" if lnk["verified"] else "no",
                        lnk["confidence"],
                    ])
            report_content = buf.getvalue()

        else:  # markdown (default)
            report_lines = []
            report_lines.append(f"# Traceability Report: {source_type}")
            report_lines.append(f"\n**Project**: {self.project}")
            report_lines.append(f"**Date**: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            report_lines.append(f"\n---\n")
            for entry in matrix_data:
                report_lines.append(f"\n## {entry['source_type']} → {entry['target_type']}")
                if not entry["found"]:
                    report_lines.append(f"\n⚠ Matrix not found: {entry['matrix']}\n")
                    continue
                report_lines.append(
                    f"\n**Coverage**: {entry['coverage']*100:.1f}%"
                    f" ({entry['total_links']} links)"
                )
                report_lines.append(f"\n| Source ID | Target ID | Link Type | Verified |\n")
                report_lines.append(f"|-----------|-----------|-----------|----------|\n")
                for lnk in entry["links"]:
                    verified_icon = "✓" if lnk["verified"] else "✗"
                    report_lines.append(
                        f"| {lnk['source_id']} | {lnk['target_id']}"
                        f" | {lnk['link_type']} | {verified_icon} |\n"
                    )
                report_lines.append("\n")
            report_content = "".join(report_lines)

        # Write to file if output specified
        if output:
            tmp = Path(str(output) + ".tmp")
            tmp.write_text(report_content)
            os.replace(tmp, output)
            print(f"✓ Report written to: {output}")
        else:
            print(report_content)
        
        return report_content
    
    # ========================================================================
    # Command: import
    # ========================================================================
    
    def import_data(self, file_path: str, matrix_type: str, format: str = None) -> TraceabilityMatrix:
        """
        Import traceability data from external file.
        
        Args:
            file_path: Input file path
            matrix_type: Matrix type (e.g., "requirements_to_architecture")
            format: Input format (csv, json, excel) - auto-detected if None
        
        Returns:
            TraceabilityMatrix with imported links
        """
        input_path = Path(file_path)
        
        if not input_path.exists():
            raise FileNotFoundError(f"Import file not found: {file_path}")
        
        # Auto-detect format
        if not format:
            format = input_path.suffix.lstrip('.')
            if format in ['xls', 'xlsx']:
                format = 'excel'
        
        # Parse matrix type to get source/target
        parts = matrix_type.split('_to_')
        if len(parts) != 2:
            raise ValueError(f"Invalid matrix type format: {matrix_type} (expected: source_to_target)")
        
        source_type, target_type = parts
        
        # Import based on format
        if format == 'csv':
            links = self._import_csv(file_path)
        elif format == 'json':
            links = self._import_json(file_path)
        elif format == 'excel':
            links = self._import_excel(file_path)
        else:
            raise ValueError(f"Unsupported import format: {format}")
        
        # Create matrix with imported links
        matrix = TraceabilityMatrix(
            name=matrix_type,
            source_type=source_type,
            target_type=target_type,
            project=self.project,
            links=links,
        )
        
        # Set output paths
        base_path = self.evidence_dir / matrix_type
        matrix.csv_path = str(base_path.with_suffix('.csv'))
        matrix.json_path = str(base_path.with_suffix('.json'))
        matrix.markdown_path = str(base_path.with_suffix('.md'))
        
        # Write to standard format
        self._write_csv(matrix)
        self._write_json(matrix)
        self._write_markdown(matrix)
        
        print(f"✓ Imported {len(links)} traceability links from {file_path}")
        print(f"  Format: {format}")
        print(f"  Matrix: {matrix_type}")
        print(f"  Output: {matrix.csv_path}")
        
        return matrix
    
    def _import_csv(self, file_path: str) -> List[TraceabilityLink]:
        """Import links from CSV file."""
        links = []
        
        with open(file_path, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                link = TraceabilityLink(
                    source_id=row.get('source_id', ''),
                    source_type=row.get('source_type', ''),
                    target_id=row.get('target_id', ''),
                    target_type=row.get('target_type', ''),
                    link_type=row.get('link_type', 'allocated_to'),
                    rationale=row.get('rationale', ''),
                    verified=row.get('verified', 'false').lower() == 'true',
                    verified_by=row.get('verified_by', ''),
                    verified_date=row.get('verified_date', ''),
                    confidence=row.get('confidence', 'high'),
                    source_document=row.get('source_document', ''),
                )
                links.append(link)
        
        return links
    
    def _import_json(self, file_path: str) -> List[TraceabilityLink]:
        """Import links from JSON file."""
        with open(file_path, 'r') as f:
            data = json.load(f)
        
        links = []
        for link_data in data.get('links', []):
            link = TraceabilityLink(
                source_id=link_data.get('source_id', ''),
                source_type=link_data.get('source_type', ''),
                target_id=link_data.get('target_id', ''),
                target_type=link_data.get('target_type', ''),
                link_type=link_data.get('link_type', 'allocated_to'),
                rationale=link_data.get('rationale', ''),
                verified=link_data.get('verified', False),
                verified_by=link_data.get('verified_by', ''),
                verified_date=link_data.get('verified_date', ''),
                confidence=link_data.get('confidence', 'high'),
                source_document=link_data.get('source_document', ''),
            )
            links.append(link)
        
        return links
    
    def _import_excel(self, file_path: str) -> List[TraceabilityLink]:
        """Import links from Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl library required for Excel import. Install: pip install openpyxl")
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        links = []
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            
            link = TraceabilityLink(
                source_id=row_dict.get('source_id', ''),
                source_type=row_dict.get('source_type', ''),
                target_id=row_dict.get('target_id', ''),
                target_type=row_dict.get('target_type', ''),
                link_type=row_dict.get('link_type', 'allocated_to'),
                rationale=row_dict.get('rationale', ''),
                verified=str(row_dict.get('verified', 'false')).lower() == 'true',
                verified_by=row_dict.get('verified_by', ''),
                verified_date=row_dict.get('verified_date', ''),
                confidence=row_dict.get('confidence', 'high'),
                source_document=row_dict.get('source_document', ''),
            )
            links.append(link)
        
        return links
    
    # ========================================================================
    # Command: export
    # ========================================================================
    
    def export_data(self, matrix_name: str, format: str, output_path: str) -> str:
        """
        Export traceability data to external format.
        
        Args:
            matrix_name: Matrix name (or "all" for all matrices)
            format: Output format (csv, json, excel, markdown)
            output_path: Output file path
        
        Returns:
            Output file path
        """
        if matrix_name == "all":
            # Export all matrices — only JSON is supported for bulk export
            if format != "json":
                print(
                    f"Error: --matrix all only supports --format json "
                    f"(got '{format}'). To export a single matrix in another "
                    f"format, specify a matrix name instead of 'all'.",
                    file=sys.stderr,
                )
                sys.exit(1)
            matrices = [p.stem for p in self.evidence_dir.glob("*.csv")]
            
            if format == "json":
                # Combine all matrices into single JSON
                all_data = []
                for name in matrices:
                    csv_path = self.evidence_dir / f"{name}.csv"
                    matrix = self._read_csv(str(csv_path))
                    all_data.append({
                        'name': matrix.name,
                        'source_type': matrix.source_type,
                        'target_type': matrix.target_type,
                        'links': [asdict(link) for link in matrix.links],
                    })
                
                with open(output_path, 'w') as f:
                    json.dump({'project': self.project, 'matrices': all_data}, f, indent=2)
                
                print(f"✓ Exported {len(matrices)} matrices to {output_path}")
            # (non-JSON for "all" is blocked above)
        
        else:
            # Export single matrix
            csv_path = self.evidence_dir / f"{matrix_name}.csv"
            
            if not csv_path.exists():
                raise FileNotFoundError(f"Matrix not found: {matrix_name}")
            
            matrix = self._read_csv(str(csv_path))
            
            if format == "csv":
                # Copy CSV
                import shutil
                shutil.copy(matrix.csv_path, output_path)
            
            elif format == "json":
                data = {
                    'name': matrix.name,
                    'source_type': matrix.source_type,
                    'target_type': matrix.target_type,
                    'project': self.project,
                    'links': [asdict(link) for link in matrix.links],
                }
                with open(output_path, 'w') as f:
                    json.dump(data, f, indent=2)
            
            elif format == "markdown":
                # Regenerate markdown from the in-memory matrix rather than
                # copying matrix.markdown_path — the .md sidecar may not exist
                # if the matrix was created or modified without a prior sync.
                self._write_markdown(matrix)
                import shutil
                shutil.copy(matrix.markdown_path, output_path)
            
            elif format == "excel":
                self._export_excel(matrix, output_path)
            
            else:
                raise ValueError(f"Unsupported export format: {format}")
            
            print(f"✓ Exported {matrix_name} to {output_path}")
            print(f"  Format: {format}")
            print(f"  Links: {len(matrix.links)}")
        
        return output_path
    
    def _export_excel(self, matrix: TraceabilityMatrix, output_path: str):
        """Export matrix to Excel format."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl library required for Excel export. Install: pip install openpyxl")
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = matrix.name[:31]  # Excel sheet name limit
        
        # Headers
        headers = ['source_id', 'source_type', 'target_id', 'target_type', 
                   'link_type', 'rationale', 'verified', 'verified_by', 'verified_date',
                   'confidence', 'source_document']
        sheet.append(headers)
        
        # Data rows
        for link in matrix.links:
            sheet.append([
                link.source_id, link.source_type, link.target_id, link.target_type,
                link.link_type, link.rationale, str(link.verified).lower(), link.verified_by,
                link.verified_date, link.confidence, link.source_document,
            ])
        
        workbook.save(output_path)
    
    # ========================================================================
    # Command: extract
    # ========================================================================
    
    def extract(self, document_path: str, link_type: str, merge: bool = False) -> List[TraceabilityLink]:
        """
        Auto-extract traceability links from documents.
        
        Args:
            document_path: Document file path (Markdown, C source, etc.)
            link_type: Link type (e.g., "design_to_requirements", "code_to_requirements")
            merge: Merge with existing matrix (default: False)
        
        Returns:
            List of extracted TraceabilityLink objects
        """
        doc_path = Path(document_path)
        
        if not doc_path.exists():
            raise FileNotFoundError(f"Document not found: {document_path}")
        
        # Read document content
        with open(doc_path, 'r') as f:
            content = f.read()
        
        # Parse link_type to determine source/target types
        # link_type format: "source_to_target" (e.g., "design_to_requirements")
        parts = link_type.split('_to_')
        if len(parts) != 2:
            raise ValueError(f"Invalid link type format: {link_type} (expected: source_to_target)")
        
        source_type, target_type = parts
        
        # Extract links using pattern matching
        extracted_links = self._extract_links_from_content(content, source_type, target_type, doc_path.name)
        
        print(f"\n{'='*70}")
        print(f"Traceability Extraction Results")
        print(f"{'='*70}")
        print(f"Document:      {document_path}")
        print(f"Link Type:     {link_type}")
        print(f"Extracted:     {len(extracted_links)} links")
        print(f"{'='*70}\n")
        
        # Show extracted links
        for link in extracted_links[:10]:  # Show first 10
            print(f"{link.source_id} → {link.target_id} (confidence: {link.confidence})")
        
        if len(extracted_links) > 10:
            print(f"... and {len(extracted_links) - 10} more")
        
        print()
        
        # Merge with existing matrix if requested
        if merge:
            matrix_name = f"{source_type}_to_{target_type}"
            csv_path = self.evidence_dir / f"{matrix_name}.csv"
            
            if csv_path.exists():
                existing_matrix = self._read_csv(str(csv_path))
                
                # Merge: keep existing links, add new ones
                existing_ids = set((link.source_id, link.target_id) for link in existing_matrix.links)
                new_links = [link for link in extracted_links 
                            if (link.source_id, link.target_id) not in existing_ids]
                
                existing_matrix.links.extend(new_links)
                
                # Update files
                self._write_csv(existing_matrix)
                self._write_json(existing_matrix)
                self._write_markdown(existing_matrix)
                
                print(f"✓ Merged {len(new_links)} new links into {matrix_name}")
                print(f"  Total links: {len(existing_matrix.links)}")
            
            else:
                # Create new matrix
                matrix = TraceabilityMatrix(
                    name=matrix_name,
                    source_type=source_type,
                    target_type=target_type,
                    project=self.project,
                    links=extracted_links,
                )
                
                base_path = self.evidence_dir / matrix_name
                matrix.csv_path = str(base_path.with_suffix('.csv'))
                matrix.json_path = str(base_path.with_suffix('.json'))
                matrix.markdown_path = str(base_path.with_suffix('.md'))
                
                self._write_csv(matrix)
                self._write_json(matrix)
                self._write_markdown(matrix)
                
                print(f"✓ Created new matrix {matrix_name} with {len(extracted_links)} links")
        
        return extracted_links
    
    def _extract_links_from_content(self, content: str, source_type: str, target_type: str, 
                                     source_document: str) -> List[TraceabilityLink]:
        """Extract traceability links from document content using pattern matching."""
        links = []
        
        # Get patterns from config
        patterns = self.config.get('artifact_patterns', {})
        
        # Determine patterns based on artifact types
        target_pattern = patterns.get(target_type)
        if not target_pattern:
            # Use generic pattern
            target_pattern = r'[A-Z]+-[A-Z]+-[A-Z0-9]+-\d+'
        
        # Find all target artifact references
        target_matches = re.finditer(target_pattern, content)
        
        # For each match, try to determine source artifact
        # Heuristic: look for section headers or explicit source IDs near the reference
        for match in target_matches:
            target_id = match.group(0)
            
            # Try to find source ID in surrounding context
            start = max(0, match.start() - 500)
            end = min(len(content), match.end() + 500)
            context = content[start:end]
            
            # Look for source artifact pattern
            source_pattern = patterns.get(source_type, r'[A-Z]+-[A-Z]+-[A-Z0-9]+-\d+')
            source_match = re.search(source_pattern, context)
            
            if source_match:
                source_id = source_match.group(0)
                
                # Determine confidence based on proximity
                distance = abs(source_match.start() - (match.start() - start))
                if distance < 100:
                    confidence = "high"
                elif distance < 300:
                    confidence = "medium"
                else:
                    confidence = "low"
                
                link = TraceabilityLink(
                    source_id=source_id,
                    source_type=source_type,
                    target_id=target_id,
                    target_type=target_type,
                    link_type="allocated_to",
                    rationale=f"Auto-extracted from {source_document}",
                    verified=False,
                    confidence=confidence,
                    source_document=source_document,
                )
                
                links.append(link)
        
        # Deduplicate links
        unique_links = {}
        for link in links:
            key = (link.source_id, link.target_id)
            if key not in unique_links:
                unique_links[key] = link
            else:
                # Keep higher confidence link
                if link.confidence == "high":
                    unique_links[key] = link
        
        return list(unique_links.values())
    
    # ========================================================================
    # Command: visualize
    # ========================================================================
    
    def visualize(self, source_type: str, target_type: str, format: str = "mermaid", 
                  output: str = None) -> str:
        """
        Generate traceability visualization.
        
        Args:
            source_type: Source artifact type or specific artifact ID
            target_type: Target artifact type
            format: Visualization format (mermaid, dot)
            output: Output file path
        
        Returns:
            Visualization content as string
        """
        # Collect all relevant matrices
        all_matrices = []
        
        # Start from source_type and trace to target_type
        # This is a simplified version - full implementation would do graph traversal
        matrix_name = f"{source_type}_to_{target_type}"
        csv_path = self.evidence_dir / f"{matrix_name}.csv"
        
        if not csv_path.exists():
            raise FileNotFoundError(f"Matrix not found: {matrix_name}")
        
        matrix = self._read_csv(str(csv_path))
        
        # Generate visualization based on format
        if format == "mermaid":
            viz_content = self._generate_mermaid(matrix)
        elif format == "dot":
            viz_content = self._generate_dot(matrix)
        else:
            raise ValueError(f"Unsupported visualization format: {format}")
        
        # Write to file if output specified
        if output:
            with open(output, 'w') as f:
                f.write(viz_content)
            print(f"✓ Visualization written to: {output}")
        else:
            print(viz_content)
        
        return viz_content
    
    def _generate_mermaid(self, matrix: TraceabilityMatrix) -> str:
        """Generate Mermaid diagram."""
        lines = []
        
        lines.append("```mermaid")
        lines.append("graph LR")
        lines.append(f"  subgraph {matrix.source_type}")
        
        # Collect unique sources
        sources = set(link.source_id for link in matrix.links)
        for source in sorted(sources):
            lines.append(f"    {source.replace('-', '_')}[\"{source}\"]")
        
        lines.append("  end")
        lines.append(f"  subgraph {matrix.target_type}")
        
        # Collect unique targets
        targets = set(link.target_id for link in matrix.links if link.target_id)
        for target in sorted(targets):
            lines.append(f"    {target.replace('-', '_')}[\"{target}\"]")
        
        lines.append("  end")
        lines.append("")
        
        # Add links
        for link in matrix.links:
            if link.target_id:
                source_clean = link.source_id.replace('-', '_')
                target_clean = link.target_id.replace('-', '_')
                lines.append(f"  {source_clean} --> {target_clean}")
        
        lines.append("```")
        
        return "\n".join(lines)
    
    def _generate_dot(self, matrix: TraceabilityMatrix) -> str:
        """Generate GraphViz DOT format."""
        lines = []
        
        lines.append("digraph traceability {")
        lines.append("  rankdir=LR;")
        lines.append("  node [shape=box];")
        lines.append("")
        
        # Source subgraph
        lines.append(f"  subgraph cluster_source {{")
        lines.append(f"    label=\"{matrix.source_type}\";")
        
        sources = set(link.source_id for link in matrix.links)
        for source in sorted(sources):
            lines.append(f"    \"{source}\";")
        
        lines.append("  }")
        lines.append("")
        
        # Target subgraph
        lines.append(f"  subgraph cluster_target {{")
        lines.append(f"    label=\"{matrix.target_type}\";")
        
        targets = set(link.target_id for link in matrix.links if link.target_id)
        for target in sorted(targets):
            lines.append(f"    \"{target}\";")
        
        lines.append("  }")
        lines.append("")
        
        # Links
        for link in matrix.links:
            if link.target_id:
                lines.append(f"  \"{link.source_id}\" -> \"{link.target_id}\";")
        
        lines.append("}")
        
        return "\n".join(lines)
    
    # ========================================================================
    # Command: sync
    # ========================================================================
    
    def sync(self, matrix_name: str) -> TraceabilityMatrix:
        """
        Synchronize CSV, JSON, and Markdown formats.
        
        Args:
            matrix_name: Matrix name (or "all" for all matrices)
        
        Returns:
            TraceabilityMatrix (or None if "all")
        """
        if matrix_name == "all":
            # Sync all matrices
            matrices = list(self.evidence_dir.glob("*.csv"))
            
            for csv_path in matrices:
                name = csv_path.stem
                self._sync_single_matrix(name)
            
            print(f"✓ Synchronized {len(matrices)} matrices")
            return None
        
        else:
            # Sync single matrix
            return self._sync_single_matrix(matrix_name)
    
    def _sync_single_matrix(self, matrix_name: str) -> TraceabilityMatrix:
        """Synchronize single matrix across all formats."""
        csv_path = self.evidence_dir / f"{matrix_name}.csv"
        
        if not csv_path.exists():
            raise FileNotFoundError(f"Matrix not found: {matrix_name}")
        
        # Read CSV (primary source)
        matrix = self._read_csv(str(csv_path))
        
        # Update JSON index
        self._write_json(matrix)
        
        # Regenerate Markdown report
        self._write_markdown(matrix)
        
        print(f"✓ Synchronized {matrix_name}")
        print(f"  CSV:      {matrix.csv_path}")
        print(f"  JSON:     {matrix.json_path}")
        print(f"  Markdown: {matrix.markdown_path}")
        
        return matrix
    
    # ========================================================================
    # CSV/JSON/Markdown I/O
    # ========================================================================
    
    def _read_csv(self, csv_path: str) -> TraceabilityMatrix:
        """Read traceability matrix from CSV file."""
        matrix_name = Path(csv_path).stem
        links = []
        
        try:
            with open(csv_path, 'r') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    link = TraceabilityLink(
                        source_id=row.get('source_id', ''),
                        source_type=row.get('source_type', ''),
                        target_id=row.get('target_id', ''),
                        target_type=row.get('target_type', ''),
                        link_type=row.get('link_type', 'allocated_to'),
                        rationale=row.get('rationale', ''),
                        verified=row.get('verified', 'false').lower() == 'true',
                        verified_by=row.get('verified_by', ''),
                        verified_date=row.get('verified_date', ''),
                        confidence=row.get('confidence', 'high'),
                        source_document=row.get('source_document', ''),
                    )
                    links.append(link)
        except FileNotFoundError:
            pass
        
        # Parse matrix name to get source/target types
        parts = matrix_name.split('_to_')
        source_type = parts[0] if len(parts) > 0 else "unknown"
        target_type = parts[1] if len(parts) > 1 else "unknown"
        
        matrix = TraceabilityMatrix(
            name=matrix_name,
            source_type=source_type,
            target_type=target_type,
            project=self.project,
            links=links,
            csv_path=csv_path,
        )
        
        # Set JSON and Markdown paths
        base_path = Path(csv_path).with_suffix('')
        matrix.json_path = str(base_path.with_suffix('.json'))
        matrix.markdown_path = str(base_path.with_suffix('.md'))
        
        return matrix
    
    def _write_csv(self, matrix: TraceabilityMatrix):
        """Write traceability matrix to CSV file (atomic write)."""
        fieldnames = [
            'source_id', 'source_type', 'target_id', 'target_type',
            'link_type', 'rationale', 'verified', 'verified_by', 'verified_date',
            'confidence', 'source_document',
        ]

        import io as _io
        buf = _io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=fieldnames)
        writer.writeheader()
        for link in matrix.links:
            writer.writerow({
                'source_id': link.source_id,
                'source_type': link.source_type,
                'target_id': link.target_id,
                'target_type': link.target_type,
                'link_type': link.link_type,
                'rationale': link.rationale,
                'verified': str(link.verified).lower(),
                'verified_by': link.verified_by,
                'verified_date': link.verified_date,
                'confidence': link.confidence,
                'source_document': link.source_document,
            })

        tmp = Path(str(matrix.csv_path) + ".tmp")
        tmp.write_text(buf.getvalue(), newline='')
        os.replace(tmp, matrix.csv_path)

    def _write_json(self, matrix: TraceabilityMatrix):
        """Write traceability matrix to JSON file (atomic write)."""
        data = {
            'name': matrix.name,
            'source_type': matrix.source_type,
            'target_type': matrix.target_type,
            'project': matrix.project,
            'links': [asdict(link) for link in matrix.links],
            'coverage': matrix.coverage,
            'version': matrix.version,
            'created_date': matrix.created_date,
            'modified_date': matrix.modified_date,
        }

        tmp = Path(str(matrix.json_path) + ".tmp")
        tmp.write_text(json.dumps(data, indent=2))
        os.replace(tmp, matrix.json_path)

    def _write_markdown(self, matrix: TraceabilityMatrix):
        """Write traceability matrix to Markdown file (atomic write)."""
        lines = []

        lines.append(f"# Traceability Matrix: {matrix.name}\n\n")
        lines.append(f"**Project**: {matrix.project}  \n")
        lines.append(f"**Source Type**: {matrix.source_type}  \n")
        lines.append(f"**Target Type**: {matrix.target_type}  \n")
        lines.append(f"**Version**: {matrix.version}  \n")
        lines.append(f"**Status**: {matrix.status}  \n")
        lines.append(f"\n---\n\n")

        if matrix.links:
            lines.append("| Source ID | Target ID | Link Type | Verified | Rationale |\n")
            lines.append("|-----------|-----------|-----------|----------|------------|\n")

            for link in matrix.links:
                verified_icon = "✓" if link.verified else "✗"
                lines.append(
                    f"| {link.source_id} | {link.target_id} | {link.link_type}"
                    f" | {verified_icon} | {link.rationale} |\n"
                )
        else:
            lines.append("*No traceability links defined yet.*\n")

        lines.append(f"\n---\n\n")
        lines.append(f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*\n")

        tmp = Path(str(matrix.markdown_path) + ".tmp")
        tmp.write_text("".join(lines))
        os.replace(tmp, matrix.markdown_path)

# ============================================================================
# CLI Interface
# ============================================================================

def main():
    """Main CLI entry point."""
    parser = argparse.ArgumentParser(
        description="Traceability Manager for EN 50128 Projects",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # Create new traceability matrix
    %(prog)s create --from requirements --to architecture
    
    # Validate traceability for SIL 3
    %(prog)s validate --phase design --sil 3
    
    # Query forward traceability
    %(prog)s query --source SW-REQ-015 --direction forward
    
    # Check for traceability gaps
    %(prog)s check-gaps --phase design --sil 3
    
    # Generate traceability report
    %(prog)s report --from requirements --to architecture,design,tests --format markdown
    
    # Import traceability from CSV
    %(prog)s import --file legacy_traces.csv --type requirements_to_design
    
    # Export traceability to JSON
    %(prog)s export --matrix all --format json --output traceability_export.json
    
    # Auto-extract traceability from document
    %(prog)s extract --document docs/Software-Design-Specification.md --type design_to_requirements --merge
    
    # Generate visualization
    %(prog)s visualize --from requirements --to architecture --format mermaid
    
    # Synchronize all formats
    %(prog)s sync --matrix all
        """
    )
    
    parser.add_argument('--project', help='Project name (overrides LIFECYCLE_STATE.md)')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
    
    subparsers = parser.add_subparsers(dest='command', help='Commands')
    
    # create command
    create_parser = subparsers.add_parser('create', help='Create new traceability matrix')
    create_parser.add_argument('--from', dest='source', required=True, help='Source artifact type')
    create_parser.add_argument('--to', dest='target', required=True, help='Target artifact type')
    create_parser.add_argument('--output', help='Output path (default: evidence/traceability/)')
    
    # validate command
    validate_parser = subparsers.add_parser('validate', help='Validate traceability completeness')
    validate_parser.add_argument('--matrix', help='Specific matrix to validate')
    validate_parser.add_argument('--phase', help='Validate all matrices for lifecycle phase')
    validate_parser.add_argument('--sil', type=int, required=True, help='SIL level (0-4)')
    
    # query command
    query_parser = subparsers.add_parser('query', help='Query traceability links')
    query_parser.add_argument('--source', help='Source artifact ID')
    query_parser.add_argument('--target', help='Target artifact ID')
    query_parser.add_argument('--direction', choices=['forward', 'backward', 'both'], default='forward', help='Query direction')
    
    # check-gaps command
    gaps_parser = subparsers.add_parser('check-gaps', help='Detect traceability gaps')
    gaps_parser.add_argument('--phase', help='Lifecycle phase to check')
    gaps_parser.add_argument('--sil', type=int, required=True, help='SIL level (0-4)')
    
    # report command
    report_parser = subparsers.add_parser('report', help='Generate traceability report')
    report_parser.add_argument('--from', dest='source', required=True, help='Source artifact type')
    report_parser.add_argument('--to', dest='targets', required=True, help='Comma-separated target types')
    report_parser.add_argument('--format', choices=['markdown', 'csv', 'json'], default='markdown', help='Output format')
    report_parser.add_argument('--output', help='Output file path')
    
    # import command
    import_parser = subparsers.add_parser('import', help='Import traceability data')
    import_parser.add_argument('--file', dest='file', required=True, help='Input file path')
    import_parser.add_argument('--type', dest='matrix_type', required=True, help='Matrix type (e.g., requirements_to_architecture)')
    import_parser.add_argument('--format', choices=['csv', 'json', 'excel'], help='Input format (auto-detected if not specified)')
    
    # export command
    export_parser = subparsers.add_parser('export', help='Export traceability data')
    export_parser.add_argument(
        '--matrix', required=True,
        help='Matrix name, or "all" to export every matrix (JSON format only when using "all")',
    )
    export_parser.add_argument('--format', choices=['csv', 'json', 'excel', 'markdown'], required=True, help='Output format (use "json" when --matrix all)')
    export_parser.add_argument('--output', required=True, help='Output file path')
    
    # extract command
    extract_parser = subparsers.add_parser('extract', help='Auto-extract traceability from documents')
    extract_parser.add_argument('--document', required=True, help='Document file path')
    extract_parser.add_argument('--type', dest='link_type', required=True, help='Link type (e.g., design_to_requirements)')
    extract_parser.add_argument('--merge', action='store_true', help='Merge with existing matrix')
    
    # visualize command
    visualize_parser = subparsers.add_parser('visualize', help='Generate traceability visualization')
    visualize_parser.add_argument('--from', dest='source', required=True, help='Source artifact type')
    visualize_parser.add_argument('--to', dest='target', required=True, help='Target artifact type')
    visualize_parser.add_argument('--format', choices=['mermaid', 'dot'], default='mermaid', help='Visualization format')
    visualize_parser.add_argument('--output', help='Output file path')
    
    # sync command
    sync_parser = subparsers.add_parser('sync', help='Synchronize CSV/JSON/Markdown formats')
    sync_parser.add_argument('--matrix', required=True, help='Matrix name (or "all" for all matrices)')

    # gate-check command
    gate_check_parser = subparsers.add_parser(
        'gate-check',
        help='Check T1-T15 normative traceability rules for a lifecycle phase gate',
    )
    gate_check_parser.add_argument(
        '--phase',
        required=True,
        choices=[
            'planning', 'requirements', 'design', 'component-design',
            'implementation-testing', 'integration', 'validation',
        ],
        help='Lifecycle phase to check (cumulative — all rules valid up to this phase)',
    )
    gate_check_parser.add_argument(
        '--sil', type=int, required=True, help='SIL level (0-4)',
    )

    # verify-link command
    verify_link_parser = subparsers.add_parser(
        'verify-link',
        help='Mark a traceability link as VER-verified in a matrix CSV',
    )
    verify_link_parser.add_argument(
        '--matrix', required=True,
        help='Matrix file stem (e.g., doc6_to_doc9)',
    )
    verify_link_parser.add_argument(
        '--source', required=True,
        help='Source artifact ID (e.g., SRS-REQ-001)',
    )
    verify_link_parser.add_argument(
        '--target', required=True,
        help='Target artifact ID (e.g., ARCH-001)',
    )
    verify_link_parser.add_argument(
        '--name', dest='verified_by', required=True,
        help='VER reviewer name or role (e.g., VER or "J.Smith")',
    )
    verify_link_parser.add_argument(
        '--date',
        help='Verification date in ISO format YYYY-MM-DD (defaults to today)',
    )
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return 1
    
    # Initialize manager
    mgr = TraceabilityManager(project=args.project)
    
    # Execute command
    try:
        if args.command == 'create':
            mgr.create_matrix(args.source, args.target, args.output)
        
        elif args.command == 'validate':
            result = mgr.validate(args.matrix, args.phase, args.sil)
            return 0 if result['overall_pass'] else 1
        
        elif args.command == 'query':
            if not args.source and not args.target:
                print("Error: query requires at least one of --source or --target.", file=sys.stderr)
                return 1
            mgr.query(args.source, args.target, args.direction)
        
        elif args.command == 'check-gaps':
            mgr.check_gaps(args.phase, args.sil)
        
        elif args.command == 'report':
            targets = args.targets.split(',')
            mgr.report(args.source, targets, args.format, args.output)
        
        elif args.command == 'import':
            mgr.import_data(args.file, args.matrix_type, args.format)
        
        elif args.command == 'export':
            mgr.export_data(args.matrix, args.format, args.output)
        
        elif args.command == 'extract':
            mgr.extract(args.document, args.link_type, args.merge)
        
        elif args.command == 'visualize':
            mgr.visualize(args.source, args.target, args.format, args.output)
        
        elif args.command == 'sync':
            mgr.sync(args.matrix)
        
        elif args.command == 'gate-check':
            result = mgr.gate_check(args.phase, args.sil)
            return 0 if result.get('overall_pass') else 1

        elif args.command == 'verify-link':
            ok = mgr.verify_link(
                args.matrix,
                args.source,
                args.target,
                args.verified_by,
                args.date,
            )
            return 0 if ok else 1
        
        return 0
    
    except Exception as e:
        print(f"✗ Error: {e}", file=sys.stderr)
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1

if __name__ == '__main__':
    sys.exit(main())
